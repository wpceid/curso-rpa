from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Carregar planilha existente
wb = load_workbook('idade_maior.xlsx')
ws = wb.active

# Acessar e modificar células
valor = ws['B2'].value
ws['C5'] = 'Atualizado'

# Formatação
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].fill = PatternFill(start_color="FFFF00", fill_type="solid")
ws['A1'].alignment = Alignment(horizontal='center')

# Criar nova planilha
nova_planilha = wb.create_sheet("Resumo")
nova_planilha['A1'] = "Relatório de Vendas"

# Salvar alterações
wb.save('relatorio_formatado.xlsx')